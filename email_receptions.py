import imaplib
import email
import openpyxl

# Login credentials
imap_host = 'imap.example.com'
imap_port = 993
username = 'user@example.com'
password = 'your_password'

# Connect to IMAP server
imap_server = imaplib.IMAP4_SSL(imap_host, imap_port)
imap_server.login(username, password)
imap_server.select('INBOX')

# Search for all email messages
status, email_ids = imap_server.search(None, 'ALL')
email_ids = email_ids[0].split()

# Create a new Excel workbook and worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = 'Email Receptions'

# Write headers to the worksheet
worksheet['A1'] = 'Email Address'
worksheet['B1'] = 'Name'

# Loop through each email message
for email_id in email_ids:
    # Fetch the email message
    status, email_data = imap_server.fetch(email_id, '(RFC822)')
    email_message = email.message_from_bytes(email_data[0][1])
    
    # Get the sender's email address and name
    sender_email = email.utils.parseaddr(email_message['From'])[1]
    sender_name = email.utils.parseaddr(email_message['From'])[0]
    
    # Write the sender's email address and name to the worksheet
    row = worksheet.max_row + 1
    worksheet.cell(row=row, column=1, value=sender_email)
    worksheet.cell(row=row, column=2, value=sender_name)

# Save the Excel workbook
workbook.save('email_receptions.xlsx')

# Disconnect from IMAP server
imap_server.close()
imap_server.logout()
